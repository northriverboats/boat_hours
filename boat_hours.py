#!/usr/bin/env python
# lag.py [cuttoffDate] [startDate]

import os
import sys
import datetime
import pymssql
import click
from dotenv import load_dotenv
from envelopes import Envelope
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from smtplib import SMTPException # allow for silent fail in try exceptio



def resource_path(relative_path: str) -> str:
    """Get absolute path to resource, works for dev and for PyInstaller

    Arguments:
        relative_path: str -- releative path from current directory + file name

    Returns:
        str -- absoulte path to file
    """
    try:
        # pylint: disable=protected-access
        base_path: str = sys._MEIPASS  # type: ignore
    except AttributeError:
        base_path: str = os.path.abspath(".")  #type: ignore

    return os.path.join(base_path, relative_path)

def get_boats(cursor):
    """Get the list of jobs 

    Aarguments:
        cursor: -- sql cursor

    Retunrs:
        list of job ids, boat names

    """
    SQL = """
    SELECT  DISTINCT(tp.job_id), job.jobname
      FROM  timeWorkingPunch tp
 LEFT JOIN  job on tp.job_id = job.job_id
     WHERE  tp.inpunch_dt > DATEADD(year,-1,GETDATE())
       AND  tp.active_yn = 1
       AND  tp.job_id > 7000
    """
    cursor.execute(SQL)
    return cursor.fetchall()

def get_boat_hours(cursor, job):
    """ get hours for boat"""
    SQL = """
        SELECT  substring(departmentname, 1,3 ) as departmentname, 
               SUM(CASE 
                 WHEN DATEPART(MINUTE, workingpunch_ts) = 45 THEN DATEPART(HOUR, workingpunch_ts) + .7
                 WHEN DATEPART(MINUTE, workingpunch_ts) = 30 THEN DATEPART(HOUR, workingpunch_ts) + .5
                 WHEN DATEPART(MINUTE, workingpunch_ts) = 15 THEN DATEPART(HOUR, workingpunch_ts) + .25
                 WHEN DATEPART(MINUTE, workingpunch_ts) =  0 THEN DATEPART(HOUR, workingpunch_ts)
               END) as WorkTime
         FROM  timeWorkingPunch tp
    LEFT JOIN  job on tp.job_id = job.job_id
    LEFT JOIN  task on tp.task_id = task.task_id
         JOIN  empMain em ON tp.employee_id = em.employee_id
         JOIN  tblDepartment dp ON tp.department_id  = dp.department_id
        WHERE  tp.job_id = %s
     GROUP BY  substring(departmentname, 1,3 )
    """
    cursor.execute(SQL, job)
    return dict(cursor.fetchall())

def get_latest_punch(cursor, job):
    SQL = """
         SELECT  max(workingpunch_ts)
           FROM  timeWorkingPunch tp
          WHERE  tp.job_id = %s
            AND  tp.active_yn = 1
            AND  tp.department_id = 221
    """
    cursor.execute(SQL, job)
    return cursor.fetchone()[0]

def valid_jobs(cursor, all_jobs):
    """ 
    latest = None
             No outfitting hours, if fab hours then ok
    latest > cutoff
             currently if fab
    """
    cutoff = datetime.datetime.today() - datetime.timedelta(days=15)
    jobs = []
    for job in all_jobs:
        latest = get_latest_punch(cursor, job[0])
        if not (latest is None or latest > cutoff):
            continue
        result = get_boat_hours(cursor, job[0])
        jobs.append([job[1], result])
    return sorted(jobs)


def nice_number(hours, dept):
    if hours.get(dept, None) is None:
        return "       "
    return f"{hours.get(dept):7.2f}"

def text_report(results):
    output = "Boat        Fabrication      Paint     Canvas Outfitting\n"
    output += "----------  -----------  ---------  --------- ----------\n"
    for result in results:
        buffer = f"{result[0]:14.12}  "
        buffer += nice_number(result[1], 'Fab')
        buffer += "    " + nice_number(result[1], 'Pai')
        buffer += "    " + nice_number(result[1], 'Can')
        buffer += "    " + nice_number(result[1], 'Out')
        output += buffer + "\n"
    return output

def write_spreadsheet(results, verbose):
    """Write results as a spreadsheet"""
    wb = Workbook()
    sh = wb.active
    bold = Font(bold=True)
    sh.column_dimensions['A'].width = 16
    sh.column_dimensions['B'].width = 12
    sh.column_dimensions['C'].width = 12
    sh.column_dimensions['D'].width = 12
    sh.column_dimensions['E'].width = 12
    sh['A1'].font = bold
    sh['B1'].font =  bold
    sh['C1'].font =  bold
    sh['D1'].font =  bold
    sh['E1'].font =  bold
    sh["A1"] = "Boat"
    sh["B1"] = "Fabrication"
    sh["C1"] = "Paint"
    sh["D1"] = "Canvas"
    sh["E1"] = "Outfitting"

    for y, result in enumerate(results, start=2):
        sh[f"A{y}"] = result[0]
        sh[f"B{y}"] = result[1].get('Fab', '')
        sh[f"C{y}"] = result[1].get('Pai', '')
        sh[f"D{y}"] = result[1].get('Can', '')
        sh[f"E{y}"] = result[1].get('Out', '')
        sh[f"B{y}"].number_format = '#,##0.00'
        sh[f"C{y}"].number_format = '#,##0.00'
        sh[f"D{y}"].number_format = '#,##0.00'
        sh[f"E{y}"].number_format = '#,##0.00'

    #Change background color of even rows
    """
    for rows in sh.iter_rows(min_row=1, max_row=1, min_col=None, max_col=5):
        for cell in rows:
            cell.fill = PatternFill(start_color="A0A0A0", end_color="A0A0A0",fill_type = "solid")
    for rows in sh.iter_rows(min_row=2, max_row=y, min_col=None, max_col=5):
        for cell in rows:
            if not cell.row % 2:
                cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0",fill_type = "solid")
    """
    sh.freeze_panes = "A2"
    sh.print_title_rows='1:1'
    file_name = datetime.datetime.today().strftime(
            os.environ.get('XLS_PATH','/tmp/') +
            os.environ.get('XLS_NAME','temp.xlsx'))
    wb.save(file_name)
    message(verbose, 1, f"Writing Spreadsheet {file_name}")

def message(verbose, limit, text):
    if verbose >= limit:
        click.echo(text)

@click.command()
@click.option('-v', '--verbose', count=True)
@click.option('-d', '--debug', is_flag=True, default=False, help="Debug send no email")
def main(verbose, debug):
    load_dotenv(dotenv_path=resource_path(".env"))
    with pymssql.connect(
            os.getenv('DB_HOST'),
            os.getenv('DB_USER'),
            os.getenv('DB_PASSWORD'),
            os.getenv('DB_DATABASE'),
            tds_version=r'7.0') as conn:
        with conn.cursor() as cursor:
            jobs = sorted(get_boats(cursor))
            results = valid_jobs(cursor, jobs)
    text_results = text_report(results)
    message(verbose, 2, text_results)
    if debug:
        print(text_results)
        message(verbose, 0, "Debug enabled, not writitng spreadsheet")
    else:
        write_spreadsheet(results, verbose)


if __name__ == "__main__":
    main()  # pylint: disable=no-value-for-parameter

